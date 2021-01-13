#!python3

import os
import statistics as st
import openpyxl as opx

from build_db import engine, User, Album, Entry, Genre
from sqlalchemy.orm import sessionmaker

entries_excelpath = "top2020_entries.xlsx"
albumlist_excelpath = "top2020_albums_list.xlsx"
genres_excelpath = "genres_list.xlsx"

Session = sessionmaker(bind=engine)
session = Session()


def add_albums():

    entries_wb = opx.load_workbook(entries_excelpath)
    entries = entries_wb['entries_linear']

    for row in entries.rows:
        
        # skip header row
        if row[0].value == 'id':
            continue
        
        entry_id = row[0].value
        album_name = row[3].value
        album_obj = session.query(Album).filter(Album.name == album_name).first()
        if album_obj is None:
            album_obj = Album(
                name = album_name
            )
            session.add(album_obj)
        
        entry_obj = session.query(Entry).filter(Entry.id == entry_id).first()
        entry_obj.album = album_obj

def update_genres_db():

    wb = opx.load_workbook(genres_excelpath)
    ws = wb['genres_list']

    for row in ws:
        genre_name = row[0].value
        genre_obj = session.query(Genre).filter(Genre.name == genre_name).first()
        if genre_obj is None:
            genre_obj = Genre()
            session.add(genre_obj)

        genre_obj.name = genre_name

def export_genres_list(excelpath):

    wb = opx.load_workbook(excelpath)
    ws = wb.create_sheet('genres_list')    
    genres_list = session.query(Genre)

    for genre in genres_list:
        ws.append([genre.name])

    wb.save(excelpath)

def export_album_list():

    albums = session.query(Album).all()

    workbook = opx.Workbook()
    ws = workbook.create_sheet(title='albums_scores')
    header = [
        'id',
        'name',
        'genre'
    ]
    ws.append(header)

    for album in albums:
        ws.append(album.get_values())

    workbook.save(albumlist_excelpath)


def main():

    add_albums()
    export_album_list()

    update_genres_db()
    
    # albums_wb = opx.load_workbook(albums_excelpath)
    export_genres_list(albumlist_excelpath)

    session.commit()
    
if __name__ == "__main__":
    main()
