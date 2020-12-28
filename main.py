#!python3

import os
import statistics as st
import openpyxl as opx

from build_db import engine, User, Album, Entry
from sqlalchemy.orm import sessionmaker

excelpath = "scoring top2020.xlsx"

Session = sessionmaker(bind=engine)
session = Session()

def load_worksheets():

    wb = opx.load_workbook(excelpath)
    scoregrid = wb['scoregrid']
    entries = wb['entries_p']

    return scoregrid, entries, wb

def add_user_entry(raw_entry):

    user_name = raw_entry[0].value
    top_size = raw_entry[1].value
    print("======================== NAME: {} ========================".format(user_name))
    
    user_obj = session.query(User).filter(User.name == user_name).first()
    if user_obj is None:
        user_obj = User(
            name = user_name,
            top_size = top_size
        )
        session.add(user_obj)

    for i in range(2,32):
        
        raw_name = raw_entry[i].value
        position = i-1
        
        if raw_name:
        
            entry_name = raw_name.lower()
            print(entry_name)

            # album_obj = session.query(Album).filter(Album.name == album_name).first()
            # if album_obj is None:
            #     album_obj = Album(
            #        name = album_name
            #     )
            #     session.add(album_obj)

            entry_score = score(position,top_size)
            print('Entry Score: {}'.format(entry_score))
            entry_obj = Entry(
                user = user_obj,
                name = entry_name,
                # album = album_obj,
                position = position,
                score = entry_score
            )
            session.add(entry_obj)

        else:
            break

def score(position, top_size):

    return 30 - 25*(position-1)/(top_size-1)

def calculate_album_stats():

    for album in session.query(Album):
        
        entries = session.query(Entry).join(Album).filter(Album.id == album.id).all()
        
        scores = set(entry.score for entry in entries)
        album.total_score = sum(scores)
        album.min_score = min(scores)
        album.max_score = max(scores)
        album.average_score = st.mean(scores)

        positions = set(entry.position for entry in entries)
        album.min_pos = min(positions)
        album.max_pos = max(positions)
        album.average_pos = st.mean(positions)

        album.nb_votes = len(entries)

        album.print_fields()

def export_album_entries(workbook):

    entries = session.query(Entry).all()
    ws = workbook.create_sheet(title='entries_linear')
    header = [
        'id',
        'user',
        'entry_name',
        'position',
        'score'
    ]
    ws.append(header)
    
    for entry in entries:
        ws.append(entry.get_values())

    workbook.save(excelpath)

def add_albums(workbook):

    ws = workbook['entries_linear']

    for row in ws.rows:
        
        if row[0].value == 'id':
            continue
        
        entry_id = row[0].value
        album_name = row[2].value
        album_obj = session.query(Album).filter(Album.name == album_name).first()
        if album_obj is None:
            album_obj = Album(
                name = album_name
            )
            session.add(album_obj)
        
        entry_obj = session.query(Entry).filter(Entry.id == entry_id).first()
        entry_obj.album = album_obj

    # session.commit()

def export_albums(workbook):

    albums = session.query(Album).all()
    ws = workbook.create_sheet(title='albums_scores')
    header = [
        'id',
        'name',
        'genre_01',
        'genre_02',
        'total_score',
        'nb_votes',
        'min_score',
        'max_score',
        'average_score',
        'min_pos',
        'max_pos',
        'average_pos'
    ]
    ws.append(header)
    
    for album in albums:
        ws.append(album.get_values())

    workbook.save(excelpath)

def main():

    scoregrid, entries, wb = load_worksheets()

    # for raw_entry in entries.rows:
    #     add_user_entry(raw_entry)

    # export_album_entries(wb)

    # add_albums(wb)
    # calculate_album_stats()

    export_albums(wb)

    session.commit()

    


if __name__ == "__main__":
    main()
