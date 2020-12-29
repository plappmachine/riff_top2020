#!python3

import os
import openpyxl as opx

from build_db import engine, User, Album, Entry
from sqlalchemy.orm import sessionmaker

input_excelpath = "top2020_entries.xlsx"
output_excelpath = "top2020_entries.xlsx"

Session = sessionmaker(bind=engine)
session = Session()

def load_worksheet():

    wb = opx.load_workbook(input_excelpath)
    entries = wb['entries_linear']

    return entries, wb

def add_albums(ws):

    for row in ws.rows:
        
        # skip header row
        if row[0].value == 'id':
            continue
        
        entry_id = row[0].value
        album_name = row[2].value
        album_obj = session.query(Album).filter(Album.name == album_name).first()
        if album_obj is None:
            album_obj = Album(
                name = album_name
            )
            session.add(album_obj)
        
        entry_obj = session.query(Entry).filter(Entry.id == entry_id).first()
        entry_obj.album = album_obj

def export_genres_list(workbook):

    genres_list = [
        'Heavy Metal',
        'Thrash Metal',
        'Prog Metal',
        'Death Metal',
        'Black Metal',
        'Folk - Pagan Metal',
        'Power Metal',
        'Post-Black Metal',
        'Avant-Garde Metal',
        'Indus Metal',
        'Doom Metal',
        'Gothic Metal',
        'Sludge',
        'Stoner - Psych',
        'Prog - Psychedelic Rock',
        'Post-Rock',
        'Post-Metal',
        'Post-Hardcore',
        'Punk',
        'HxC',
        'Crust',
        'Grind',
        'Metalcore - Mathcore',
        'Deathcore',
        'Experimental',
        'Goth Rock',
        'Shoegaze',
        'Post-Punk',
        'Industrial',
        'Neofolk - DarkFolk',
        'Pop Rock - Indie Rock',
        'Blues - Blues Rock',
        'Funk Rock',
        'Funk Metal',
        'Grunge',
        'Alternative Rock / Metal',
        'Neo Metal',
        'Rap Metal',
        'Hard Rock',
        "Metal'n'Roll"
    ]
    ws = None
    try:
        ws = workbook['genres_list']
    except Exception as e:
        ws = workbook.create_sheet('genres_list')
    
    for genre in genres_list:
        ws.append([genre])

    workbook.save(output_excelpath)

def update_albums_genres(ws):

    for row in ws.rows:
        
        album_id = row[0].value
        genre = row[2].value
        # skip header row
        if album_id  == 'id':
            continue
        
        album_obj = session.query(Album).filter(Album.id == album_id).first()
        album_obj.genre = genre

def calculate_album_stats():

    for album in session.query(Album):
        
        entries = session.query(Entry).join(Album).filter(Album.id == album.id).all()
        
        scores = set(entry.score for entry in entries)
        album.total_score = sum(scores)
        album.min_score = min(scores)
        album.max_score = max(scores)
        album.average_score = st.mean(scores)

        positions = set(entry.position for entry in entries)
        album.min_pos = min(positions)
        album.max_pos = max(positions)
        album.average_pos = st.mean(positions)

        album.nb_votes = len(entries)

        album.print_fields()

def export_albums(workbook):

    albums = session.query(Album).all()
    
    ws = None
    try:
        ws = workbook['albums_scores']
    except Exception as e:
        ws = workbook.create_sheet(title='albums_scores')
    
    header = [
        'id',
        'name',
        'genre',
        'total_score',
        'nb_votes',
        'min_score',
        'max_score',
        'average_score',
        'min_pos',
        'max_pos',
        'average_pos'
    ]
    ws.append(header)
    
    for album in albums:
        ws.append(album.get_values())

    workbook.save(output_excelpath)

def main():

    entries, wb = load_worksheet()
    
    add_albums(entries)
    calculate_album_stats()
    export_albums(wb)
    export_genres_list(wb)

    session.commit()
    
if __name__ == "__main__":
    main()