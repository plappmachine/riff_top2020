#!python3

import os
import openpyxl as opx

from build_db import engine, User, Album, Entry
from sqlalchemy.orm import sessionmaker

input_excelpath = "scoring top2020.xlsx"
output_excelpath = "top2020_entries.xlsx"

Session = sessionmaker(bind=engine)
session = Session()

def load_worksheet():

    wb = opx.load_workbook(input_excelpath)
    entries = wb['entries']

    return entries, wb

def add_user_entry(raw_entry):

    user_name = raw_entry[1].value
    top_size = len(list(x for x in raw_entry if x.value is not None))-2

    print("======================== NAME: {} ========================".format(user_name))
    print('TOP SIZE = {}'.format(top_size))
    
    # create User db object if does not already exist
    user_obj = session.query(User).filter(User.name == user_name).first()
    if user_obj is None:
        user_obj = User(
            name = user_name,
            top_size = top_size
        )
        session.add(user_obj)

    for i in range(2,top_size+2):
        
        raw_name = raw_entry[i].value
        entry_name = raw_name.lower()
        # print(entry_name)
        position = i-1

        entry_score = score(position,top_size)
        print('Entry Score: {}'.format(entry_score))
        
        # create Entry db object if does not already exist
        entry_obj = session.query(Entry).filter(Entry.name == entry_name).first()
        if entry_obj is None:
            entry_obj = Entry(
                user = user_obj,
                name = entry_name
            )
            session.add(entry_obj)
        
        entry_obj.position = position
        entry_obj.score = entry_score


def score(position, top_size):

    return 30 - 25*(position-1)/(top_size-1)

def export_linear_entries():

    wb = opx.Workbook()
    entries = session.query(Entry).all()
    ws = wb.create_sheet(title='entries_linear')
    header = [
        'id',
        'user',
        'entry_name',
        'position',
        'score'
    ]
    ws.append(header)
    
    for entry in entries:
        ws.append(entry.get_values())

    wb.save(output_excelpath)

def main():

    entries, wb = load_worksheet()

    for raw_entry in entries.rows:
        add_user_entry(raw_entry)

    export_linear_entries()

    session.commit()


if __name__ == "__main__":
    main()
